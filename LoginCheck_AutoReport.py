import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

REPORT_FILE = "Reporte_LOGIN.xlsx"

COLOR_HEADER_BG   = "2F5496"
COLOR_PASSED_BG   = "E2EFDA"
COLOR_FAILED_BG   = "FFDCE1"
COLOR_HEADER_FONT = "FFFFFF"

def export_result_to_xlsx(test_name, status, details):
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    new_row = {
        "Fecha y Hora":    timestamp,
        "Caso de Prueba":  test_name,
        "Resultado":       status,
        "Detalles / Error": details
    }

    if os.path.isfile(REPORT_FILE):
        df_existing = pd.read_excel(REPORT_FILE)
        df = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)
    else:
        df = pd.DataFrame([new_row])

    df.to_excel(REPORT_FILE, index=False, sheet_name="Reporte")

    wb = load_workbook(REPORT_FILE)
    ws = wb["Reporte"]

    col_widths = {"A": 22, "B": 35, "C": 12, "D": 60}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font      = Font(bold=True, color=COLOR_HEADER_FONT, name="Arial", size=11)
        cell.fill      = PatternFill("solid", start_color=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[1].height = 20

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_val = row[2].value
        row_color  = COLOR_PASSED_BG if status_val == "PASSED" else COLOR_FAILED_BG
        for cell in row:
            cell.fill      = PatternFill("solid", start_color=row_color)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = border

    ws.freeze_panes = "A2"
    wb.save(REPORT_FILE)

def run_test():
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    wait = WebDriverWait(driver, 10)
    test_name = "Detailed Login Health Check"
    
    # Rastreo de errores
    errores = []
    
    try:
        driver.get("https://karaokeivan.com/index.php/login/")
        driver.maximize_window()

        # Validar Usuario
        try:
            user_input = wait.until(EC.presence_of_element_located((By.ID, "user_login")))
            user_input.send_keys("FakeUser")
        except:
            errores.append("Usuario (user_login)")

        # Validar Clave
        try:
            pass_input = driver.find_element(By.ID, "user_pass") 
            pass_input.send_keys("FakePassword")
        except:
            errores.append("Clave (user_pass)")

        # Validar Boton
        try:
            boton = driver.find_element(By.CSS_SELECTOR, "input[type='submit']")
        except:
            errores.append("Boton Login")

        # Evaluacion final
        if not errores:
            export_result_to_xlsx(test_name, "PASSED", "Todos los elementos (Usuario, Clave, Boton) localizados.")
            print("✅ Test PASSED")
        else:
            cantidad = len(errores)
            detalle_error = f"Fallaron {cantidad} elementos: {', '.join(errores)}"
            export_result_to_xlsx(test_name, "FAILED", detalle_error)
            print(f"❌ Test FAILED: {detalle_error}")

    except Exception as e:
        export_result_to_xlsx(test_name, "CRITICAL", f"Error de carga de pagina: {str(e)[:30]}")
    
    finally:
        driver.quit()

if __name__ == "__main__":
    run_test()